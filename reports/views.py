import csv
import io
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.db.models import Count, F, Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views import View
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from tasks.models import Task

User = get_user_model()


def _date_filter(range_val):
    today = timezone.now().date()
    if range_val == '7':
        return {'created_at__date__gte': today - timedelta(days=7)}
    if range_val == '30':
        return {'created_at__date__gte': today - timedelta(days=30)}
    if range_val == 'month':
        return {'created_at__year': today.year, 'created_at__month': today.month}
    return {}


def _task_counts(qs):
    today = timezone.now().date()
    return qs.aggregate(
        total=Count('id'),
        done=Count('id', filter=Q(status='done')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        todo=Count('id', filter=Q(status='todo')),
        overdue=Count('id', filter=Q(
            status__in=['todo', 'in_progress'],
            due_date__isnull=False,
            due_date__lt=today,
        )),
    )


class MyPerformanceView(APIView):
    """
    GET /api/v1/reports/my-performance/
    Employees see own stats. Managers pass ?user_id=<id> for a direct report.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        subject = request.user

        if request.user.role in ['manager', 'admin']:
            user_id = request.query_params.get('user_id', '').strip()
            if user_id:
                try:
                    target = User.objects.get(id=int(user_id), is_active=True)
                except (User.DoesNotExist, ValueError):
                    return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
                if request.user.role == 'manager' and target.manager_id != request.user.id:
                    return Response(
                        {'error': 'You can only view performance for your direct reports.'},
                        status=status.HTTP_403_FORBIDDEN,
                    )
                subject = target

        qs = Task.objects.filter(assigned_to=subject, is_deleted=False)

        stats = qs.aggregate(
            total_tasks=Count('id'),
            done_tasks=Count('id', filter=Q(status='done')),
            on_time_count=Count(
                'id',
                filter=Q(
                    status='done',
                    done_at__isnull=False,
                    due_date__isnull=False,
                    done_at__date__lte=F('due_date'),
                ),
            ),
            late_count=Count(
                'id',
                filter=Q(
                    status='done',
                    done_at__isnull=False,
                    due_date__isnull=False,
                    done_at__date__gt=F('due_date'),
                ),
            ),
        )

        total = stats['total_tasks']
        done  = stats['done_tasks']
        completion_rate = round(done / total * 100, 1) if total > 0 else 0.0

        return Response({
            'total_tasks':     total,
            'done_tasks':      done,
            'completion_rate': completion_rate,
            'on_time_count':   stats['on_time_count'],
            'late_count':      stats['late_count'],
            'user': {
                'id':        subject.id,
                'full_name': f'{subject.first_name} {subject.last_name}'.strip() or subject.username,
                'role':      subject.role,
            },
        })


class TeamSummaryView(APIView):
    """
    GET /api/v1/reports/team-summary/?range=all|7|30|month
    Managers see their direct reports. Admins see all employees.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role not in ['manager', 'admin']:
            return Response({'error': 'Not authorised.'}, status=status.HTTP_403_FORBIDDEN)

        range_val   = request.query_params.get('range', 'all')
        df          = _date_filter(range_val)

        employees_qs = User.objects.filter(role='employee', is_active=True)
        base_qs      = Task.objects.filter(is_deleted=False, **df)

        if request.user.role == 'manager':
            employees_qs = employees_qs.filter(manager=request.user)
            team_ids     = list(employees_qs.values_list('id', flat=True)) + [request.user.id]
            base_qs      = base_qs.filter(assigned_to__in=team_ids)

        agg = _task_counts(base_qs)
        t, d = agg['total'] or 0, agg['done'] or 0

        team = []
        for emp in employees_qs.order_by('first_name', 'last_name'):
            ea  = _task_counts(Task.objects.filter(assigned_to=emp, is_deleted=False, **df))
            et, ed = ea['total'] or 0, ea['done'] or 0
            team.append({
                'id':              emp.id,
                'name':            f'{emp.first_name} {emp.last_name}'.strip() or emp.username,
                'department':      emp.department or '—',
                'total':           et,
                'done':            ed,
                'in_progress':     ea['in_progress'] or 0,
                'todo':            ea['todo'] or 0,
                'overdue':         ea['overdue'] or 0,
                'completion_rate': round(ed / et * 100, 1) if et > 0 else 0.0,
            })

        return Response({
            'range': range_val,
            'summary': {
                'total':           t,
                'done':            d,
                'in_progress':     agg['in_progress'] or 0,
                'todo':            agg['todo'] or 0,
                'overdue':         agg['overdue'] or 0,
                'completion_rate': round(d / t * 100, 1) if t > 0 else 0.0,
            },
            'team': team,
        })


class ExportView(APIView):
    """
    GET /api/v1/reports/export/?format=csv|excel&type=team|my_tasks&range=all|7|30|month
    Streams a file download. 'team' type requires manager/admin role.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fmt         = request.query_params.get('format', 'csv')
        export_type = request.query_params.get('type', 'my_tasks')
        range_val   = request.query_params.get('range', 'all')
        df          = _date_filter(range_val)

        if export_type == 'team':
            if request.user.role not in ['manager', 'admin']:
                return Response({'error': 'Not authorised.'}, status=status.HTTP_403_FORBIDDEN)
            headers, rows, filename = self._team_data(request.user, df)
        else:
            headers, rows, filename = self._my_tasks_data(request.user, df)

        if fmt == 'excel':
            return self._excel(headers, rows, filename)
        return self._csv(headers, rows, filename)

    def _team_data(self, user, df):
        employees = User.objects.filter(role='employee', is_active=True)
        if user.role == 'manager':
            employees = employees.filter(manager=user)

        headers = [
            'Employee', 'Department', 'Total Tasks', 'Done',
            'In Progress', 'To Do', 'Overdue', 'Completion Rate (%)',
        ]
        rows = []
        for emp in employees.order_by('first_name', 'last_name'):
            ea = _task_counts(Task.objects.filter(assigned_to=emp, is_deleted=False, **df))
            et, ed = ea['total'] or 0, ea['done'] or 0
            rows.append([
                f'{emp.first_name} {emp.last_name}'.strip() or emp.username,
                emp.department or '—',
                et, ed,
                ea['in_progress'] or 0,
                ea['todo'] or 0,
                ea['overdue'] or 0,
                round(ed / et * 100, 1) if et > 0 else 0.0,
            ])
        return headers, rows, 'team_performance_report'

    def _my_tasks_data(self, user, df):
        tasks = Task.objects.filter(
            assigned_to=user, is_deleted=False, **df
        ).order_by('-created_at')

        STATUS_MAP = {'todo': 'To Do', 'in_progress': 'In Progress', 'done': 'Done'}
        headers = ['Title', 'Description', 'Priority', 'Status', 'Due Date', 'Completed At', 'Created At']
        rows = []
        for task in tasks:
            rows.append([
                task.title,
                task.description or '',
                task.priority.capitalize(),
                STATUS_MAP.get(task.status, task.status),
                str(task.due_date) if task.due_date else '—',
                str(task.done_at.date()) if task.done_at else '—',
                str(task.created_at.date()),
            ])
        return headers, rows, 'my_tasks_report'

    def _csv(self, headers, rows, filename):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    def _excel(self, headers, rows, filename):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        ws.append(headers)
        fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            width = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return resp


class ReportsPageView(View):
    def get(self, request):
        if not request.COOKIES.get('access_token'):
            return redirect('/')
        return render(request, 'reports.html')
